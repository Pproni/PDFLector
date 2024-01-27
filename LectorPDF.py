from spire.pdf.common import *
from spire.pdf import *
import pdfplumber
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import re
import os
import numpy as np
from datetime import datetime, timedelta
import Interface

def folder_creator(folder_directory, folder_name):
    path = os.path.join(folder_directory, folder_name)
    os.makedirs(path, exist_ok=True)

def check_pdfs(path):
    files = []
    files_names = []
    for file in os.listdir(path):
        filename, extension = os.path.splitext(file)
        if extension.lower() == '.pdf':
            files.append(file)
            files_names.append(filename)
    return files, files_names

def pdf2svg(pdf,pdf_name,x=1800,y=1600):
    
    path_pdf = os.path.join(str(os.getcwd()),'PDFs')
    path_svg = os.path.join(str(os.getcwd()),'SVGs')
    path_data = os.path.join(str(os.getcwd()),'data')
    
    # Create an object of the PdfDocument class
    doc = PdfDocument()

    # Load a PDF file
    doc.LoadFromFile(os.path.join(path_pdf,str(pdf)))

    # Specify the width and height of output SVG files
    doc.ConvertOptions.SetPdfToSvgOptions(float(x), float(y))

    # Save each page of the file to a separate SVG file
    doc.SaveToFile(os.path.join(path_svg,str(pdf_name))+".svg", 0, 0, FileFormat.SVG)

    # Create a PdfDocument object
    doc1 = PdfDocument()
    # Load an SVG file
    doc1.LoadFromSvg(os.path.join(path_svg,str(pdf_name))+".svg")

    # Save the SVG file to PDF format
    doc1.SaveToFile(os.path.join(path_data,str(pdf_name))+".pdf", FileFormat.PDF)

    # Close the PdfDocument object
    doc.Close()
    doc1.Close()
    
def pdf2excel(pdf_file,pdf_name):
    path_data = os.path.join(str(os.getcwd()),'data')
    path_excel = os.path.join(str(os.getcwd()),'Excel')
    
    pdf = pdfplumber.open(os.path.join(path_data,str(pdf_file)))
    p0 = pdf.pages[0]
    table = p0.extract_table()

    # Crear DataFrame
    df = pd.DataFrame(table[1:])

    # Invertir las filas
    df = df[::-1].reset_index(drop=True)

    # Guardar en Excel
    df.to_excel(os.path.join(path_excel,str(pdf_name))+".xlsx", index=False)

def data_extractor(excel_file):
    path_excel = os.path.join(str(os.getcwd()),'Excel')
    workbook = load_workbook(filename=os.path.join(path_excel,str(excel_file))+".xlsx")
    sheet = workbook.active
    #Extracción de nombres, horas totales, hora de inicio, final, y descanso
    names = []
    total_hours = []
    time_list = []
    
    for i in range(14):
        j = 2*i+17
        cells = sheet[str("H"+str(j)): str("J"+str(j))]
        if str(sheet[str("B"+str(j))].value).replace(" ","") == str(707):
            #Nombres
            names.append(sheet[str("C"+str(j))].value)
            #Horas totales
            total_hours.append((sheet[str("L"+str(j-1))].value).replace(" ",""))
            #Horas iniciales, descanso y finales
            if str(sheet[str("I"+str(j))].value).replace(" ","") == str(0):
                clock_list = [time_list.append([str(c1.value),str(0) ,str(c2.value),str(c3.value)]) for (c1,c2,c3) in cells if str(c1.value) != 'None']
                time_list = [[h1.replace(" ",""),str(h2),h3.replace(" ",""),h4.replace(" ","")] for x in time_list for (h1,h2,h3,h4) in [x]]
                time_list = list(filter(None,time_list))                
            elif str(sheet[str("I"+str(j))].value).replace(" ","") != str(0):
                clock_list = [time_list.append([str(c1.value),str(sheet[str("I"+str(j-1))].value) ,str(c2.value),str(c3.value)]) for (c1,c2,c3) in cells if str(c1.value) != 'None']
                time_list = [[h1.replace(" ",""),str(h2).replace(" ",""),h3.replace(" ",""),h4.replace(" ","")] for x in time_list for (h1,h2,h3,h4) in [x]]
                time_list = list(filter(None,time_list))
            else:
                print('Hay un dato atípico en la fila: '+j)
                pass
    #Conversión de horas string a horas flotantes
    time_list = [[int(time) for time in sublist] for sublist in time_list]

    #Extracción de Trabajo, código de trabajo, fecha de inicio y final
    job_ID_text = re.split('(\d+)',str(sheet[str("F4")].value).replace(" ",""))
    job_name = (str(sheet[str("F7")].value).replace(" ","")).replace("\n","")
    start_day = (str(sheet[str("B5")].value).replace(" ","")).replace("STARTDATE","").replace("\n","")
    end_day = (str(sheet[str("B8")].value).replace(" ","")).replace("STOPDATE","").replace("\n","")
    total_hours = [float(x) for x in total_hours]
    names = [x.replace(" ", "") for x in names]
    for i in job_ID_text:
        if i.isdigit():
            job_ID = i        
    return names, job_name.replace("JOBNAME",""), job_ID, total_hours, time_list, start_day, end_day

def daylist_generator(initial_date, final_date):
    # Convierte las cadenas de fecha en objetos datetime
    initial_date = datetime.strptime(initial_date, "%Y-%m-%d")
    final_date = datetime.strptime(final_date, "%Y-%m-%d")

    # Calcula la diferencia entre las fechas
    diff = final_date - initial_date

    # Crea una lista de días
    day_list = [initial_date + timedelta(days=d) for d in range(diff.days + 1)]
    day_list = [(str(dia.strftime('%A'))+" "+ str(dia.day)) for dia in day_list]

    return day_list

def excel_creator(name_list, date_list, start_day, end_day):
    
    def get_excel_column_name(column_number):
        """
        Convierte el número de columna al estilo de nombre de columna de excel.
        Ej: 1->A, 2->B, 26->Z, 27->AA, etc.
        """
        result = ""
        while column_number > 0:
            remainder = (column_number - 1) % 26
            result = chr(65 + remainder) + result
            column_number = (column_number - 1) // 26
        return str(result)
    
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    #Estilo de celdas
    Titulos_fechas_style = PatternFill(fill_type="solid", fgColor="ffd966")
    Nombres_style = PatternFill(fill_type="solid", fgColor="b7b7b7")
    
    # Crear un libro de trabajo
    workbook = Workbook()

    # Crear una hoja con título específico
    #workbook.create_sheet(title='General')
    sheet = workbook['Sheet']
    sheet.title = 'General'
    
    #Función para obtener el título con los números cardinales
    def get_ordinal_suffix(day):
        if 10 <= day % 100 <= 20:
            suffix = 'th'
        else:
            suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(day % 10, 'th')
        return suffix
    
    start_date = start_day
    end_date = end_day

    formatted_start = f"{start_date.day}{get_ordinal_suffix(start_date.day)} {start_date.strftime('%b')}"
    formatted_end = f"{end_date.day}{get_ordinal_suffix(end_date.day)} {end_date.strftime('%b')}"
    
    #Colorear
    sheet['A1'].fill = PatternFill(fill_type="solid", fgColor="666666")
    sheet['L1'].fill = PatternFill(fill_type="solid", fgColor="666666")
    
    #Combinar celdas
    sheet['B1'] = f'{formatted_start} to {formatted_end} - Chicago'
    sheet['B1'].fill = PatternFill(fill_type="solid", fgColor="666666")
    sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=11)
    sheet['N1'].fill = Titulos_fechas_style
    sheet.merge_cells(start_row=1, start_column=14, end_row=1, end_column=20)
    sheet['V1'].fill = Titulos_fechas_style
    sheet.merge_cells(start_row=1, start_column=22, end_row=1, end_column=28)
    sheet['AD1'].fill = Titulos_fechas_style
    sheet.merge_cells(start_row=1, start_column=30, end_row=1, end_column=36)
    
    #Agregar nombres
    espacio_pagos = 1
    
    sheet['A2'] = 'NAMES'
    sheet['A2'].fill = Titulos_fechas_style
    sheet[get_excel_column_name(36 + espacio_pagos) + '2'] = 'NAMES'
    sheet[get_excel_column_name(36 + espacio_pagos) + '2'].fill = Titulos_fechas_style
    for i in range(len(name_list)):
        sheet['A'+str(i+3)] = name_list[i]
        sheet['A'+str(i+3)].fill = Nombres_style
        sheet[get_excel_column_name(36 + espacio_pagos) + str(i+3)] = name_list[i]
        sheet[get_excel_column_name(36 + espacio_pagos) + str(i+3)].fill = Nombres_style
    
    #Agregar los días de trabajo
    for i in range(len(date_list)):
        sheet[str(chr(66+i))+'2'] = date_list[i]
        sheet[str(chr(66+i))+'2'].fill = Titulos_fechas_style
        
    #Agregar celdas con horas diarias
    cell_text_perday = ['TOTAL HOURS DAY - DAILY', 'TOTAL REGULAR HOURS - DAILY', 'TOTAL OVERTIME HOURS - DAILY']
    titulos_totales_style = ['b6d7a8','f9cb9c','a4c2f4','ea9999']

    for i in range(len(cell_text_perday)):
        sheet['A'+str(len(name_list)+3+i)] = cell_text_perday[i]
        sheet['A'+str(len(name_list)+3+i)].fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[i])
        for j in range(len(date_list)):
            sheet[str(chr(66+j))+str(len(name_list)+3+i)].fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[i])
            if i==0:
                sheet[str(chr(66+j))+str(len(name_list)+3+i)] = "=SUM("+str(chr(66+j))+'3:'+str(chr(66+j))+str(len(name_list)+2)+')'
            elif i==1:
                sheet[str(chr(66+j))+str(len(name_list)+3+i)] = f"={get_excel_column_name(21+espacio_pagos+j)}{len(name_list)+3}"
            elif i==2:
                sheet[str(chr(66+j))+str(len(name_list)+3+i)] = f"={get_excel_column_name(29+espacio_pagos+j)}{len(name_list)+3}"
    
    #Formato y código de las horas semanales            
    cell_text_week = ['TOTAL HOURS - WEEKLY', 'TOTAL REGULAR HOURS - WEEKLY', 'TOTAL OVERTIME HOURS - WEEKLY','PAGOS']
    for i in range(len(cell_text_week)):
        sheet[str(chr(66+len(date_list)+i))+'2'] = cell_text_week[i]
        sheet[str(chr(66+len(date_list)+i))+'2'].fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[i])            
        for j in range(len(name_list)):
            sheet[str(chr(66+len(date_list)+i))+str(3+j)].fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[i])
            if i==0:
                sheet[str(chr(66+len(date_list)+i))+str(3+j)] = "=SUM("+str(chr(66))+str(3+j)+':'+str(chr(66+len(date_list)-1))+str(3+j)+')'
            elif i==1:
                sheet[str(chr(66+len(date_list)+i))+str(3+j)] = "=IF(I"+str(3+j)+'<=40,I'+str(3+j)+',40)'
            elif i==2:
                sheet[str(chr(66+len(date_list)+i))+str(3+j)] = "=I"+str(3+j)+"-J"+str(3+j)
            elif i==3:
                sheet[str(chr(66+len(date_list)+i))+str(3+j)] = f"=I{3+j}*15"
    
    #Realiza la suma de cada una de las columnas
    sheet[f"I{len(name_list)+3}"] = '=SUM(I3:I'+str(len(name_list)+2)+')'
    sheet[f"I{len(name_list)+3}"].fill = PatternFill(fill_type="solid", fgColor='93c47d')
    sheet[f"J{len(name_list)+4}"] = '=SUM(J3:J'+str(len(name_list)+2)+')'
    sheet[f"J{len(name_list)+4}"].fill = PatternFill(fill_type="solid", fgColor='f6b26b')
    sheet[f"K{len(name_list)+5}"] = '=SUM(K3:K'+str(len(name_list)+2)+')'
    sheet[f"K{len(name_list)+5}"].fill = PatternFill(fill_type="solid", fgColor='6d9eeb')
    for i in range(3,6):
        sheet[f"L{len(name_list)+i}"] = f"={get_excel_column_name(i+6)}{len(name_list)+i}*15"
        sheet[f"L{len(name_list)+i}"].fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[3])
    
    #Llena las celdas vacías
    for columna in range(9,12):
        for fila in range(int(len(name_list)+3),int(len(name_list)+6)):                
            if sheet.cell(row=fila, column=columna).value == None:
                sheet.cell(row=fila, column=columna, value='-')
                if columna == 9:
                    sheet.cell(row=fila, column=columna).fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[fila-len(name_list)-3])
                elif columna == 10:
                    sheet.cell(row=(len(name_list)+3), column=columna).fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[1])
                    sheet.cell(row=int(len(name_list)+5), column=columna).fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[2])
                elif columna == 11:
                    sheet.cell(row=fila, column=columna).fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[2])
            else:
                pass
            
    #Tablas extra
    ##Total Hours table 
    ###Nombres de las tablas
    sheet[str(chr(77+espacio_pagos))+'1'] = 'TOTAL HOURS (ACCUMULATED)'
    sheet[get_excel_column_name(21+espacio_pagos)+'1'] = 'REGULAR HOURS'
    sheet[get_excel_column_name(29+espacio_pagos)+'1'] = 'OVERTIME HOURS (PER DAY)'
    
    ###Total hours (Accumulated)
    for i in range(len(date_list)):
        sheet[str(chr(77+i+espacio_pagos))+'2'] = date_list[i]
        sheet[str(chr(77+i+espacio_pagos))+'2'].fill = Titulos_fechas_style
        for j in range(len(name_list)):
            sheet[str(chr(77+i+espacio_pagos))+str(j+3)].fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[0])
            if i==0:
                sheet[str(chr(77+i+espacio_pagos))+str(j+3)] = f"=B{j + 3}"
            else:
                previous_cell = str(chr(77+i+espacio_pagos-1))+str(j+3)
                general_table_cell = str(chr(66+i))+str(j+3)
                sheet[str(chr(77+i+espacio_pagos))+str(j+3)] = f"={general_table_cell}+{previous_cell}"
    
    ##Regular Hours
    for i in range(len(date_list)):
        sheet[get_excel_column_name(21 + i + espacio_pagos) + '2'] = date_list[i]
        sheet[get_excel_column_name(21 + i + espacio_pagos) + '2'].fill = Titulos_fechas_style
        for j in range(len(name_list)):
            sheet[get_excel_column_name(21 + i + espacio_pagos) + str(j + 3)].fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[1])
            if i == 0:
                sheet[get_excel_column_name(21 + i + espacio_pagos) + str(j + 3)] = f"=N{j + 3}"
            else:
                sheet[get_excel_column_name(21 + i + espacio_pagos) + str(j + 3)] = f"=IF({get_excel_column_name(13 + espacio_pagos + i)}{j+3}<=0, 0, IF({get_excel_column_name(13 + espacio_pagos + i)}{j+3}<=40,{get_excel_column_name(13 + espacio_pagos + i)}{j+3}-{get_excel_column_name(13 + espacio_pagos + i - 1)}{j+3},IF({get_excel_column_name(13 + espacio_pagos + i)}{j+3}-{get_excel_column_name(13 + espacio_pagos + i-1)}{j+3}<=0, 0, ABS({get_excel_column_name(13 + espacio_pagos + i)}{j+3}-{get_excel_column_name(13 + espacio_pagos + i-1)}{j+3}-{get_excel_column_name(29 + espacio_pagos + i)}{j+3}))))"

    ##Overtime Hours
    for i in range(len(date_list)):
        sheet[get_excel_column_name(29 + i + espacio_pagos) + '2'] = date_list[i]
        sheet[get_excel_column_name(29 + i + espacio_pagos) + '2'].fill = Titulos_fechas_style
        for j in range(len(name_list)):
            sheet[get_excel_column_name(29 + i + espacio_pagos) + str(j + 3)].fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[2])
            if i == 0:
                sheet[get_excel_column_name(29 + i + espacio_pagos) + str(j + 3)] = "=0"
            else:
                sheet[get_excel_column_name(29 + i + espacio_pagos) + str(j + 3)] =f"=IF({get_excel_column_name(13 + espacio_pagos + i)}{j+3}<=0, 0, IF({get_excel_column_name(13 + espacio_pagos + i)}{j+3}<=40,0, IF({get_excel_column_name(13 + espacio_pagos + i)}{j+3}-{get_excel_column_name(13 + espacio_pagos + i-1)}{j+3}<=0,0,IF({get_excel_column_name(13 + espacio_pagos + i)}{j+3}>40, {get_excel_column_name(13 + espacio_pagos + i)}{j+3}-40-SUM({get_excel_column_name(29 + espacio_pagos)}{j+3}:{get_excel_column_name(29 + espacio_pagos + i - 1)}{j+3}),0))))"
    
    for i in range(len(date_list)):
        sheet[chr(77+i+espacio_pagos)+str(len(name_list)+3)] = f"=SUM({chr(77+i+espacio_pagos)}3:{chr(77+i+espacio_pagos)}{len(name_list)+2})"
        sheet[chr(77+i+espacio_pagos)+str(len(name_list)+3)].fill = PatternFill(fill_type="solid", fgColor='93c47d')
        sheet[get_excel_column_name(21 + i + espacio_pagos) + str(len(name_list) + 3)] = f"=SUM({get_excel_column_name(21 + i + espacio_pagos)}3:{get_excel_column_name(21 + i + espacio_pagos)}{len(name_list) + 2})"
        sheet[get_excel_column_name(21 + i + espacio_pagos) + str(len(name_list) + 3)].fill = PatternFill(fill_type="solid", fgColor='f6b26b')
        sheet[get_excel_column_name(29 + i + espacio_pagos) + str(len(name_list) + 3)] = f"=SUM({get_excel_column_name(29 + i + espacio_pagos)}3:{get_excel_column_name(29 + i + espacio_pagos)}{len(name_list) + 2})"
        sheet[get_excel_column_name(29 + i + espacio_pagos) + str(len(name_list) + 3)].fill = PatternFill(fill_type="solid", fgColor='6d9eeb')
       
    # Crear un estilo para centrar el contenido
    centro_style = Alignment(horizontal="center", vertical="center",wrap_text=True)

    tot_rows = sheet.max_row #get max row number
    tot_cols = sheet.max_column #get max column number
    
    sheet.column_dimensions['A'].width = 23
    sheet.column_dimensions[get_excel_column_name(37)].width = 23
    sheet.row_dimensions[1].height = 56.25
    sheet.row_dimensions[2].height = 56.25
    
    for i in range(len(name_list)):
        sheet.row_dimensions[i+3].height = 15.75
    
    for i in range(len(name_list)+3,len(name_list)+6):
        sheet.row_dimensions[i].height = 33
    
    for i in range(2,37):
        sheet.column_dimensions[get_excel_column_name(i)].width = 11.86
    
    for c in range(1,tot_cols+1):
        for r in range(1,tot_rows+1):
            sheet.cell(row=r, column=c).alignment = centro_style
            sheet.cell(row=r, column=c).font = Font(name='Arial', size=10)
            if sheet.cell(row=r, column=c).value != None:
                sheet.cell(row=r, column=c).border = thin_border
            else:
                pass
    for c in range(2,37):
        for r in range(3,len(name_list)+6):
            sheet.cell(row=r, column=c).number_format = '0.00'
            
    sheet.cell(row=1, column=2).font = Font(name='Arial', size=24, bold=True, color='FFFFFF')
    #Fijar columnas y filas
    sheet.freeze_panes = 'B3'
    
    # Guardar el libro de trabajo en un archivo
    workbook.save('test_savedata.xlsx')
    workbook.close()

def new_sheets(name_list, job_name, job_ID, date_list, job_day, job_names, hours):
    workbook = load_workbook('test_savedata.xlsx')
    if not str(job_name) in workbook.sheetnames:
        new_sheet = workbook.create_sheet(str(job_name))
    else:
        new_sheet = workbook[str(job_name)]
    
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    #Estilo de celdas
    Titulos_fechas_style = PatternFill(fill_type="solid", fgColor="ffd966")
    Nombres_style = PatternFill(fill_type="solid", fgColor="b7b7b7")
    
    #Casillas de información
    new_sheet['B1'] = str(job_name) +'-'+str(job_ID)
    new_sheet['B1'].fill = PatternFill(fill_type="solid", fgColor="666666")
    new_sheet.merge_cells(start_row=1, start_column=2, end_row=1, end_column=11)
    new_sheet['A2'] = 'Names'
    new_sheet['A2'].fill = Titulos_fechas_style
    
    #Colorear
    new_sheet['A1'].fill = PatternFill(fill_type="solid", fgColor="666666")    
    
    #Agregar Nombres
    for i in range(len(name_list)):
        new_sheet['A'+str(i+3)] = name_list[i]
        new_sheet['A'+str(i+3)].fill = Nombres_style
    
    #Agregar fechas
    for i in range(len(date_list)):
        new_sheet[str(chr(66+i))+'2'] = date_list[i]
        new_sheet[str(chr(66+i))+'2'].fill = Titulos_fechas_style
        
    #Agregar horas
    job_day = datetime.strptime(job_day, "%m/%d/%y")
    job_day = (str(job_day.strftime('%A'))+" "+ str(job_day.day))
    
    #Cambio de los nombres a los nuevos editados
    #for i in range(len(cambios)):
    #    if cambios[i][0] in job_names:
    #        indice = job_names.index(cambios[i][0])
    #        job_names[indice] = cambios[i][1]
    #    else:
    #        pass
    
    for i in job_names:
        if i in name_list and job_day in date_list:
            row = name_list.index(str(i))
            column = date_list.index(str(job_day))
            if new_sheet[str(chr(66+int(column)))+str(int(row)+3)].value == None:
                new_sheet[str(chr(66+int(column)))+str(int(row)+3)] = '='+str("{:.2f}".format(hours[job_names.index(str(i))]))
            else:
                new_sheet[str(chr(66+int(column)))+str(int(row)+3)] = new_sheet[str(chr(66+int(column)))+str(int(row)+3)].value +'+'+str("{:.2f}".format(hours[job_names.index(str(i))]))
        else:
            print('No coincide la fecha:', job_day, ', en el trabajo: ', job_name)
            pass
    
    #Llenar espacios en blanco de las horas con ceros
    for fila in range(3, int(len(name_list)+3)):
        for columna in range(2,9):
            new_sheet.cell(row=fila, column=columna).fill = PatternFill(fill_type="solid", fgColor='b4a7d6')
            if new_sheet.cell(row=fila, column=columna).value == None:
                new_sheet.cell(row=fila, column=columna, value='=0')
            else:
                pass
    
    #Agregar celdas con horas diarias
    cell_text_perday = ['TOTAL HOURS DAY - DAILY', 'TOTAL REGULAR HOURS - DAILY', 'TOTAL OVERTIME HOURS - DAILY']
    titulos_totales_style = ['b6d7a8','f9cb9c','a4c2f4']
    for i in range(len(cell_text_perday)):
        new_sheet['A'+str(len(name_list)+3+i)] = cell_text_perday[i]
        new_sheet['A'+str(len(name_list)+3+i)].fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[i])
        #Para este caso se tiene que editar pensando en que las celdas tienen valores dados otras tablas según la plantilla que se usa generalmente, es decir,
        #los valores de la casilla Total Regular Hours - Daily debe variar si hay overtime en alguna de las casillas
        for j in range(len(date_list)):
            new_sheet[str(chr(66+j))+str(len(name_list)+3+i)].fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[i])
            if i==0:
                new_sheet[str(chr(66+j))+str(len(name_list)+3+i)] = "=SUM("+str(chr(66+j))+'3:'+str(chr(66+j))+str(len(name_list)+2)+')'
            elif i==1:
                new_sheet[str(chr(66+j))+str(len(name_list)+3+i)] = "="+str(chr(66+j))+str(len(name_list)+3)+"-"+str(chr(66+j))+str(len(name_list)+5)
            elif i==2:
                new_sheet[str(chr(66+j))+str(len(name_list)+3+i)] = "=0"
    
    #Formato y código de las horas semanales            
    cell_text_week = ['TOTAL HOURS - WEEKLY', 'TOTAL REGULAR HOURS - WEEKLY', 'TOTAL OVERTIME HOURS - WEEKLY']
    for i in range(len(cell_text_week)):
        new_sheet[str(chr(66+len(date_list)+i))+'2'] = cell_text_week[i]
        new_sheet[str(chr(66+len(date_list)+i))+'2'].fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[i])
        for j in range(len(name_list)):
            new_sheet[str(chr(66+len(date_list)+i))+str(3+j)].fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[i])
            if i==0:
                new_sheet[str(chr(66+len(date_list)+i))+str(3+j)] = "=SUM("+str(chr(66))+str(3+j)+':'+str(chr(66+len(date_list)-1))+str(3+j)+')'
            elif i==1:
                new_sheet[str(chr(66+len(date_list)+i))+str(3+j)] = "=I"+str(3+j)+"-K"+str(3+j)
            elif i==2:
                new_sheet[str(chr(66+len(date_list)+i))+str(3+j)] = "=0"
    
    #Realiza la suma de cada una de las columnas
    new_sheet[f"I{len(name_list)+3}"] = '=SUM(I3:I'+str(len(name_list)+2)+')'
    new_sheet[f"I{len(name_list)+3}"].fill = PatternFill(fill_type="solid", fgColor='93c47d')
    new_sheet[f"J{len(name_list)+4}"] = '=SUM(J3:J'+str(len(name_list)+2)+')'
    new_sheet[f"J{len(name_list)+4}"].fill = PatternFill(fill_type="solid", fgColor='f6b26b')
    new_sheet[f"K{len(name_list)+5}"] = '=SUM(K3:K'+str(len(name_list)+2)+')'
    new_sheet[f"K{len(name_list)+5}"].fill = PatternFill(fill_type="solid", fgColor='6d9eeb')
    
    #Llena las celdas vacías
    for columna in range(9,12):
        for fila in range(int(len(name_list)+3),int(len(name_list)+6)):                
            if new_sheet.cell(row=fila, column=columna).value == None:
                new_sheet.cell(row=fila, column=columna, value='-')
                if columna == 9:
                    new_sheet.cell(row=fila, column=columna).fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[fila-len(name_list)-3])
                elif columna == 10:
                    new_sheet.cell(row=(len(name_list)+3), column=columna).fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[1])
                    new_sheet.cell(row=int(len(name_list)+5), column=columna).fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[2])
                elif columna == 11:
                    new_sheet.cell(row=fila, column=columna).fill = PatternFill(fill_type="solid", fgColor=titulos_totales_style[2])
            else:
                pass
    
    centro_style = Alignment(horizontal="center", vertical="center",wrap_text=True)

    tot_rows = new_sheet.max_row #get max row number
    tot_cols = new_sheet.max_column #get max column number
    
    new_sheet.column_dimensions['A'].width = 23
    new_sheet.row_dimensions[1].height = 56.25
    new_sheet.row_dimensions[2].height = 56.25
    
    for i in range(len(name_list)):
        new_sheet.row_dimensions[i+3].height = 15.75
        
    for i in range(len(name_list)+3,len(name_list)+6):
        new_sheet.row_dimensions[i].height = 33
    for i in range(0,12):
        new_sheet.column_dimensions[chr(i+66)].width = 11.86
        
    for c in range(1,tot_cols+1):
        for r in range(1,tot_rows+1):
            new_sheet.cell(row=r, column=c).alignment = centro_style
            new_sheet.cell(row=r, column=c).font = Font(name='Arial', size=10)
            if new_sheet.cell(row=r, column=c).value != None:
                new_sheet.cell(row=r, column=c).border = thin_border
            else:
                pass
    for c in range(2,12):
        for r in range(3,len(name_list)+6):
            new_sheet.cell(row=r, column=c).number_format = '0.00'
    new_sheet.cell(row=1, column=2).font = Font(name='Arial', size=24, bold=True, color='FFFFFF')
    
    #Fijar filas y columnas
    new_sheet.freeze_panes = 'B3'
    
    #Guardado y cerrado del documento
    workbook.save('test_savedata.xlsx')
    workbook.close()

def general_hours(name_list, date_list):
    workbook = load_workbook('test_savedata.xlsx')
    
    sheet_names = workbook.sheetnames
    sheet_names = [sheet_names[i+1] for i in range(len(sheet_names)-1)]
    sheet = workbook['General']
    
    horas_style = PatternFill(fill_type="solid", fgColor='b4a7d6')
    
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    
    for i in range(len(name_list)):
        for j in range(len(date_list)):
            formula_suma = '+'.join([f"'{hoja}'!{chr(66 + int(j))}{int(i) + 3}" for hoja in sheet_names])
            sheet[str(chr(66+int(j)))+str(int(i)+3)] = f"=SUM({formula_suma})"
            sheet[str(chr(66+int(j)))+str(int(i)+3)].fill = horas_style
            sheet[str(chr(66+int(j)))+str(int(i)+3)].border = thin_border
    
    workbook.save('test_savedata.xlsx')
    workbook.close()

def Cortafuegos(name_list,job_list,hours_perjob,start_day_perjob,start_day_factura):
    workbook = load_workbook('test_savedata.xlsx')
    General_sheet = workbook['General']
    
    #Días
    days = [(start_day_factura + timedelta(days=x)) for x in range(7)]
    days = [str(x.day) for x in days]
    
    #Colores
    Colors = ['00B050','00B0F0','FFC000','CC66FF','8ED7DD','FABF8F','66FFFF']
    
    #Bordes
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))    
    
    #Títulos
    General_sheet[f"B{len(name_list)+7}"] = f"# Timesheet"
    General_sheet[f"C{len(name_list)+7}"] = f"Día"
    General_sheet[f"D{len(name_list)+7}"] = f"Horas"
    General_sheet[f"E{len(name_list)+7}"] = f"Trabajo"
    
    days_dict = {}
    for i in days:
        days_dict[str(i)] = []
    
    for i in range(len(job_list)):
        General_sheet[f"B{len(name_list)+8+i}"] = f"#{i+1}"
        General_sheet[f"C{len(name_list)+8+i}"] = f"{start_day_perjob[i]}"
        General_sheet[f"D{len(name_list)+8+i}"] = f"={sum(hours_perjob[i])}"
        General_sheet[f"D{len(name_list)+8+i}"].number_format = '0.00'
        General_sheet[f"E{len(name_list)+8+i}"] = f"{job_list[i]}"
        if str(General_sheet[f"C{len(name_list)+8+i}"].value) in days:
            index = days.index(General_sheet[f"C{len(name_list)+8+i}"].value)
            General_sheet[f"C{len(name_list)+8+i}"].fill = PatternFill(fill_type="solid", fgColor=Colors[index])
            days_dict[str(General_sheet[f"C{len(name_list)+8+i}"].value)].append(sum(hours_perjob[i]))
        else:
            print('pasa algo')
            
    centro_style = Alignment(horizontal="center", vertical="center",wrap_text=True)
                
    for i in range(len(days)):
        General_sheet[f"F{len(name_list)+8+i}"] = f"={sum(days_dict[str(days[i])])}"
        General_sheet[f"F{len(name_list)+8+i}"].number_format = '0.00'
        General_sheet[f"F{len(name_list)+8+i}"].fill = PatternFill(fill_type="solid", fgColor=Colors[i])
        General_sheet[f"F{len(name_list)+8+i}"].border = thin_border
        General_sheet[f"F{len(name_list)+8+i}"].alignment = centro_style        
    General_sheet[f"F{len(name_list)+8+len(days)}"] = f"=SUM(F{len(name_list)+8}:F{len(name_list)+7+len(days)})"
    General_sheet[f"F{len(name_list)+8+len(days)}"].fill = PatternFill(fill_type="solid", fgColor='FFFF00')
    General_sheet[f"F{len(name_list)+8+len(days)}"].number_format = '0.00'
    General_sheet[f"F{len(name_list)+8+len(days)}"].border = thin_border
    General_sheet[f"F{len(name_list)+8+len(days)}"].alignment = centro_style  
    
    #Parte final
    General_sheet[f"C{len(name_list)+8+len(job_list)}"] = f"TOTAL"
    General_sheet[f"C{len(name_list)+8+len(job_list)}"].fill = PatternFill(fill_type="solid", fgColor='FFFF00')
    General_sheet[f"D{len(name_list)+8+len(job_list)}"] = f"=SUM(D{len(name_list)+8}:D{len(name_list)+7+len(job_list)})"
    General_sheet[f"D{len(name_list)+8+len(job_list)}"].number_format = '0.00'
    General_sheet[f"D{len(name_list)+8+len(job_list)}"].fill = PatternFill(fill_type="solid", fgColor='FFFF00')
    
    for c in range(2,7):
        for r in range(len(name_list)+7,len(name_list)+7+len(job_list)+3):
            General_sheet.cell(row=r, column=c).alignment = centro_style
            General_sheet.cell(row=r, column=c).font = Font(name='Arial', size=10)
            if General_sheet.cell(row=r, column=c).value != None:
                General_sheet.cell(row=r, column=c).border = thin_border
            else:
                pass            
            
    workbook.save('test_savedata.xlsx')
    workbook.close()